from aiogram import Router, F
from aiogram.types import Message, FSInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from database import get_conn
import config
import io

router = Router()
SUPER_ADMIN_ID = 6302553503


def is_super(telegram_id):
    return telegram_id == SUPER_ADMIN_ID


def super_menu():
    from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="📊 Статистика"), KeyboardButton(text="👥 Все компании")],
        [KeyboardButton(text="🚫 Заблокировать"), KeyboardButton(text="✅ Разблокировать")],
        [KeyboardButton(text="🗑 Удалить из ЧС"), KeyboardButton(text="🗑 Удалить товар")],
        [KeyboardButton(text="📤 Экспорт в Excel"), KeyboardButton(text="📊 Топ компаний")],
        [KeyboardButton(text="📡 Отправить всё в канал")],,
        [KeyboardButton(text="🏠 Главное меню")],
    ], resize_keyboard=True)


class SuperStates(StatesGroup):
    block_id = State()
    unblock_id = State()
    delete_bl_id = State()
    delete_prod_id = State()


@router.message(F.text == "👑 Супер Админ")
async def super_admin_menu(msg: Message):
    if not is_super(msg.from_user.id):
        return
    await msg.answer("👑 <b>Панель супер-администратора</b>", parse_mode="HTML", reply_markup=super_menu())


@router.message(F.text == "📊 Статистика")
async def stats(msg: Message):
    if not is_super(msg.from_user.id):
        return
    conn = get_conn()
    users = conn.execute("SELECT COUNT(*) FROM users WHERE status='approved'").fetchone()[0]
    pending = conn.execute("SELECT COUNT(*) FROM users WHERE status='pending'").fetchone()[0]
    blocked = conn.execute("SELECT COUNT(*) FROM users WHERE status='blocked'").fetchone()[0]
    bl = conn.execute("SELECT COUNT(*) FROM blacklist").fetchone()[0]
    prods = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    conn.close()
    await msg.answer(
        f"📊 <b>Статистика</b>\n\n"
        f"✅ Активных компаний: {users}\n"
        f"⏳ Ожидают одобрения: {pending}\n"
        f"🚫 Заблокированных: {blocked}\n"
        f"🔴 Записей в ЧС: {bl}\n"
        f"📦 Товаров в базе: {prods}",
        parse_mode="HTML"
    )


@router.message(F.text == "📊 Топ компаний")
async def top_companies(msg: Message):
    if not is_super(msg.from_user.id):
        return
    conn = get_conn()
    bl_top = conn.execute("""
        SELECT added_by_company, added_by_name, COUNT(*) as cnt
        FROM blacklist
        WHERE added_by_company != ''
        GROUP BY added_by_company
        ORDER BY cnt DESC LIMIT 10
    """).fetchall()
    prod_top = conn.execute("""
        SELECT added_by_company, added_by_name, COUNT(*) as cnt
        FROM products
        WHERE added_by_company != ''
        GROUP BY added_by_company
        ORDER BY cnt DESC LIMIT 10
    """).fetchall()
    conn.close()

    text = "📊 <b>Топ компаний по активности</b>\n\n"
    text += "🚫 <b>По чёрному списку:</b>\n"
    if bl_top:
        for i, r in enumerate(bl_top, 1):
            text += f"{i}. {r['added_by_company']} — {r['cnt']} записей\n"
    else:
        text += "Нет данных\n"

    text += "\n📦 <b>По базе товаров:</b>\n"
    if prod_top:
        for i, r in enumerate(prod_top, 1):
            text += f"{i}. {r['added_by_company']} — {r['cnt']} товаров\n"
    else:
        text += "Нет данных\n"

    await msg.answer(text, parse_mode="HTML")


@router.message(F.text == "📤 Экспорт в Excel")
async def export_excel(msg: Message):
    if not is_super(msg.from_user.id):
        return
    await msg.answer("⏳ Генерирую Excel файл...")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import tempfile, os

        wb = openpyxl.Workbook()

        # ── Лист 1: Чёрный список ──
        ws1 = wb.active
        ws1.title = "Чёрный список"
        headers1 = ["ID", "ФИО", "Дата рождения", "Паспорт", "ПИНФЛ", "Адрес",
                    "Телефон", "Причина", "Компания", "Сотрудник", "Тел. сотрудника", "Дата"]
        ws1.append(headers1)
        for cell in ws1[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="C00000")
            cell.alignment = Alignment(horizontal="center")

        conn = get_conn()
        bl_rows = conn.execute("SELECT * FROM blacklist ORDER BY id DESC").fetchall()
        for r in bl_rows:
            ws1.append([
                r["id"], r["full_name"], r["birth_date"], r["passport_number"],
                r["pinfl"], r["address"], r["phone"], r["reason"],
                r["added_by_company"], r["added_by_name"], r["added_by_phone"],
                r["added_at"][:10] if r["added_at"] else ""
            ])

        # ── Лист 2: База товаров ──
        ws2 = wb.create_sheet("База товаров")
        headers2 = ["ID", "Серийный номер", "Товар", "Покупатель", "Телефон",
                    "Паспорт", "ПИНФЛ", "Сумма", "Оплачено", "Остаток",
                    "Компания", "Сотрудник", "Дата"]
        ws2.append(headers2)
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F497D")
            cell.alignment = Alignment(horizontal="center")

        prod_rows = conn.execute("SELECT * FROM products ORDER BY id DESC").fetchall()
        for r in prod_rows:
            ws2.append([
                r["id"], r["serial_number"], r["product_name"], r["buyer_name"],
                r["buyer_phone"], r["buyer_passport"], r["buyer_pinfl"],
                r["total_amount"], r["paid_amount"],
                (r["total_amount"] or 0) - (r["paid_amount"] or 0),
                r["added_by_company"], r["added_by_name"],
                r["added_at"][:10] if r["added_at"] else ""
            ])
        conn.close()

        # Авторазмер колонок
        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        # Сохраняем во временный файл
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        tmp.close()

        await msg.answer_document(
            FSInputFile(tmp.name, filename="nasiya_baza_export.xlsx"),
            caption=f"📊 Экспорт базы данных\n🔴 ЧС: {len(bl_rows)} записей\n📦 Товары: {len(prod_rows)} записей"
        )
        os.unlink(tmp.name)

    except Exception as e:
        await msg.answer(f"❌ Ошибка: {e}")


@router.message(F.text == "👥 Все компании")
async def list_users(msg: Message):
    if not is_super(msg.from_user.id):
        return
    conn = get_conn()
    users = conn.execute("SELECT * FROM users ORDER BY registered_at DESC").fetchall()
    conn.close()
    if not users:
        await msg.answer("Пользователей нет.")
        return
    text = "👥 <b>Все компании:</b>\n\n"
    for u in users:
        status_emoji = {"approved": "✅", "pending": "⏳", "rejected": "❌", "blocked": "🚫"}.get(u["status"], "❓")
        text += (
            f"{status_emoji} <b>{u['full_name']}</b>\n"
            f"🏢 {u['company'] or '—'}\n"
            f"📱 {u['phone']}\n"
            f"🆔 <code>{u['telegram_id']}</code>\n"
            f"{'─' * 20}\n"
        )
    await msg.answer(text, parse_mode="HTML")


@router.message(F.text == "🚫 Заблокировать")
async def block_start(msg: Message, state: FSMContext):
    if not is_super(msg.from_user.id):
        return
    await msg.answer("🆔 Введите Telegram ID компании для блокировки:")
    await state.set_state(SuperStates.block_id)


@router.message(SuperStates.block_id)
async def block_do(msg: Message, state: FSMContext):
    await state.clear()
    try:
        user_id = int(msg.text.strip())
    except ValueError:
        await msg.answer("❌ Неверный ID", reply_markup=super_menu())
        return
    conn = get_conn()
    user = conn.execute("SELECT * FROM users WHERE telegram_id=?", (user_id,)).fetchone()
    conn.execute("UPDATE users SET status='blocked' WHERE telegram_id=?", (user_id,))
    conn.commit()
    conn.close()
    name = user["full_name"] if user else str(user_id)
    await msg.answer(f"🚫 <b>{name}</b> заблокирован.", parse_mode="HTML", reply_markup=super_menu())
    try:
        await msg.bot.send_message(user_id, "🚫 Ваш аккаунт заблокирован администратором системы.")
    except Exception:
        pass


@router.message(F.text == "✅ Разблокировать")
async def unblock_start(msg: Message, state: FSMContext):
    if not is_super(msg.from_user.id):
        return
    await msg.answer("🆔 Введите Telegram ID компании для разблокировки:")
    await state.set_state(SuperStates.unblock_id)


@router.message(SuperStates.unblock_id)
async def unblock_do(msg: Message, state: FSMContext):
    await state.clear()
    try:
        user_id = int(msg.text.strip())
    except ValueError:
        await msg.answer("❌ Неверный ID", reply_markup=super_menu())
        return
    conn = get_conn()
    user = conn.execute("SELECT * FROM users WHERE telegram_id=?", (user_id,)).fetchone()
    conn.execute("UPDATE users SET status='approved' WHERE telegram_id=?", (user_id,))
    conn.commit()
    conn.close()
    name = user["full_name"] if user else str(user_id)
    await msg.answer(f"✅ <b>{name}</b> разблокирован.", parse_mode="HTML", reply_markup=super_menu())
    try:
        await msg.bot.send_message(user_id, "✅ Ваш аккаунт разблокирован.")
    except Exception:
        pass


@router.message(F.text == "🗑 Удалить из ЧС")
async def del_bl_start(msg: Message, state: FSMContext):
    if not is_super(msg.from_user.id):
        return
    conn = get_conn()
    rows = conn.execute("SELECT id, full_name, added_at FROM blacklist ORDER BY id DESC LIMIT 20").fetchall()
    conn.close()
    if not rows:
        await msg.answer("ЧС пуст.", reply_markup=super_menu())
        return
    text = "🗑 <b>Последние записи в ЧС:</b>\n\n"
    for r in rows:
        text += f"#{r['id']} — {r['full_name']} ({r['added_at'][:10]})\n"
    text += "\nВведите ID записи для удаления:"
    await msg.answer(text, parse_mode="HTML")
    await state.set_state(SuperStates.delete_bl_id)


@router.message(SuperStates.delete_bl_id)
async def del_bl_do(msg: Message, state: FSMContext):
    await state.clear()
    try:
        record_id = int(msg.text.strip())
    except ValueError:
        await msg.answer("❌ Неверный ID", reply_markup=super_menu())
        return
    conn = get_conn()
    conn.execute("DELETE FROM blacklist WHERE id=?", (record_id,))
    conn.commit()
    conn.close()
    await msg.answer(f"✅ Запись #{record_id} удалена из ЧС.", reply_markup=super_menu())


@router.message(F.text == "🗑 Удалить товар")
async def del_prod_start(msg: Message, state: FSMContext):
    if not is_super(msg.from_user.id):
        return
    conn = get_conn()
    rows = conn.execute("SELECT id, product_name, serial_number FROM products ORDER BY id DESC LIMIT 20").fetchall()
    conn.close()
    if not rows:
        await msg.answer("База товаров пуста.", reply_markup=super_menu())
        return
    text = "🗑 <b>Последние товары:</b>\n\n"
    for r in rows:
        text += f"#{r['id']} — {r['product_name']} ({r['serial_number']})\n"
    text += "\nВведите ID товара для удаления:"
    await msg.answer(text, parse_mode="HTML")
    await state.set_state(SuperStates.delete_prod_id)


@router.message(SuperStates.delete_prod_id)
async def del_prod_do(msg: Message, state: FSMContext):
    await state.clear()
    try:
        record_id = int(msg.text.strip())
    except ValueError:
        await msg.answer("❌ Неверный ID", reply_markup=super_menu())
        return
    conn = get_conn()
    conn.execute("DELETE FROM products WHERE id=?", (record_id,))
    conn.commit()
    conn.close()
    await msg.answer(f"✅ Товар #{record_id} удалён.", reply_markup=super_menu())


@router.message(F.text == "📡 Отправить всё в канал")
async def send_all_to_channel(msg: Message):
    if not is_super(msg.from_user.id):
        return
    conn = get_conn()
    rows = conn.execute("SELECT * FROM blacklist ORDER BY id DESC").fetchall()
    conn.close()

    if not rows:
        await msg.answer("ЧС пуст.", reply_markup=super_menu())
        return

    await msg.answer(f"⏳ Отправляю {len(rows)} записей в канал...")

    import config
    sent = 0
    for r in rows:
        text = (
            f"🚨 <b>Чёрный список</b>\n\n"
            f"👤 <b>{r['full_name']}</b>\n"
            f"📅 {r['birth_date'] or '—'}\n"
            f"📄 {r['passport_number'] or '—'}\n"
            f"🔢 {r['pinfl'] or '—'}\n"
            f"🏠 {r['address'] or '—'}\n"
            f"📱 {r['phone'] or '—'}\n"
            f"📝 {r['reason']}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"🏢 {r['added_by_company'] or '—'}\n"
            f"👤 {r['added_by_name'] or '—'}\n"
            f"📅 {r['added_at'][:10]}"
        )
        try:
            await msg.bot.send_message(config.CHANNEL_ID, text, parse_mode="HTML")
            sent += 1
        except Exception as e:
            print(f"Ошибка: {e}")

    await msg.answer(f"✅ Отправлено {sent} из {len(rows)} записей в канал!", reply_markup=super_menu())
